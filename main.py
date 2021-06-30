'''
Copyright [2020] <Copyright yourui>
Author: yourui
Date: 2021-06-28
'''
from openpyxl import Workbook
import pickle

import requests  # 导入requests包
from bs4 import BeautifulSoup
import os

url = 'https://www.qingting.fm'


def getProvinceRadio(link, province_name):
    temp_radio_list = []
    total_page = 1000
    # province_all = requests.get(link)
    # province_all_soup = BeautifulSoup(province_all.text, 'lxml')

    for page in range(1, total_page + 1):
        province_page_temp = requests.get(link+'/' + str(page))
        province_page_soup = BeautifulSoup(province_page_temp.text, 'lxml')
        radiodata = province_page_soup.select(
            '#app > div > div.p-radio-page-root > div > div.bodyContainer > div.c-radio-page-body-root > div > div.contentSec > div > a ')
        if len(radiodata) == 0:
            break
        for list in range(1, len(radiodata), 2):
            result = {
                'id': radiodata[list].get('href').split('/')[2],
                'title': radiodata[list].get_text(),
                'province': province_name
            }
            temp_radio_list.append(result)
    return temp_radio_list


def outputXML(radio_list):

    wb = Workbook()
    dest_filename = 'QingtingFM.xlsx'
    ws1 = wb.active  # 第一个表
    ws1.title = '蜻蜓FM电台'
    ws1['A1'] = '电台名'
    ws1['B1'] = '省份'
    ws1['C1'] = 'URL'
    for row in range(2, len(radio_list) + 2):
        ws1.cell(row, 1, radio_list[row - 2].get('title'))
        ws1.cell(row, 2, radio_list[row - 2].get('province'))
        ws1.cell(row, 3, 'https://lhttp.qingting.fm/live/' +
                 radio_list[row-2].get('id') + '/64k.mp3')
    wb.save(filename=dest_filename)  # 保存


def outpitStream(radio_list):
    # 打开一个文件
    fo = open("live_streams.sii", "w", encoding='utf_8_sig')
    fo.write('SiiNunit\n{\nlive_stream_def : _nameless.1FDE.8F10 {\n stream_data: ' +
             str(len(radio_list)) + '\n')  # \n表示换行
    for item in range(0, len(radio_list)):
        fo.write(' stream_data['+str(item)+']: "' + 'https://lhttp.qingting.fm/live/' +
                 radio_list[item].get('id') + '/64k.mp3|'+radio_list[item].get('title').replace("\"", " ")+'|'+radio_list[item].get("province")+'|CN|128|0"\n')
    fo.write('}\n\n}\n')  # \n表示换行
    fo.close()


def main():
    radio_list = []
    if os.path.exists('radio_list.pickle'):
        print("本地存在文件")
        with open('radio_list.pickle', 'rb') as f:
            radio_list = pickle.load(f)
        outputXML(radio_list)
        outpitStream(radio_list)
        return
    print("本地不存在文件")
    # 访问首页获取省份信息
    strhtml = requests.get(url+'/radiopage')
    soup = BeautifulSoup(strhtml.text, 'lxml')
    provincedata = soup.select(
        '#app > div > div.p-radio-page-root > div > div.bodyContainer > div.c-radio-page-body-root > div > div.catSec > div.channelMenu.regionChoose.focusMenu > div.regionsSec.regionsSecHide > a')

    # 省份信息存储在pro_result_list
    pro_result_list = []
    for item in provincedata:
        result = {
            'province': item.get_text(),
            'id': item.get('id'),
            'link': url + item.get('href')
        }
        pro_result_list.append(result)
    pro_result_list.append({
        'province':'国家台',
        'id':'409',
        'link':'/radiopage/409'
    })
    pro_result_list.append({
        'province':'网络台',
        'id':'407',
        'link':'/radiopage/407'
    })
        # 电台信息存储在radio_list
    for privince_info in pro_result_list:
        temp_url = url+'/radiopage/' + privince_info.get('id')
        print('正在获取:'+privince_info.get('province')+'\t电台信息' +
              '\t网页地址为: ' + temp_url)
        radio_list.extend(getProvinceRadio(
            temp_url, privince_info.get('province')))
    with open('radio_list.pickle', 'wb') as f:
        pickle.dump(radio_list, f, pickle.HIGHEST_PROTOCOL)

    outputXML(radio_list)
    outpitStream(radio_list)


if (__name__ == "__main__"):
    main()
